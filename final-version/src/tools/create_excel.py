"""创建试题Excel文件"""
import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_questions_excel(questions_file: str, output_file: str):
    """
    创建试题Excel文件
    
    Args:
        questions_file: JSON题目文件路径
        output_file: Excel输出文件路径
    """
    # 加载题目数据
    with open(questions_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # 创建工作簿
    wb = Workbook()
    
    # 定义样式
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 难度颜色
    difficulty_colors = {
        '基础': PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'),  # 绿色
        '中等': PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'),  # 橙色
        '困难': PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid'),  # 红色
    }
    
    content_font = Font(name='微软雅黑', size=11)
    content_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== 单选题工作表 ==========
    ws_single = wb.active
    ws_single.title = "单选题"
    
    # 设置表头
    single_headers = ['题号', '章节', '章节名称', '页码', '题目内容', '选项A', '选项B', '选项C', '选项D', '答案', '解析', '难度']
    ws_single.append(single_headers)
    
    # 设置表头样式
    for col_num, header in enumerate(single_headers, 1):
        cell = ws_single.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 填充数据
    for idx, question in enumerate(data['single_choice'], 1):
        options = question.get('options', [])
        
        row_data = [
            idx,
            question.get('chapter', ''),
            question.get('chapter_title', ''),
            question.get('page', ''),
            question.get('question', ''),
            options[0].replace('A. ', '') if len(options) > 0 else '',
            options[1].replace('B. ', '') if len(options) > 1 else '',
            options[2].replace('C. ', '') if len(options) > 2 else '',
            options[3].replace('D. ', '') if len(options) > 3 else '',
            question.get('answer', ''),
            question.get('analysis', ''),
            question.get('difficulty', '中等')
        ]
        
        ws_single.append(row_data)
        
        # 设置内容样式
        row_num = idx + 1
        difficulty = question.get('difficulty', '中等')
        difficulty_fill = difficulty_colors.get(difficulty, difficulty_colors['中等'])
        
        for col_num in range(1, len(row_data) + 1):
            cell = ws_single.cell(row=row_num, column=col_num)
            cell.font = content_font
            cell.alignment = content_alignment
            cell.border = thin_border
            
            # 为整行设置难度颜色
            if col_num <= 3:  # 前三列（题号、章节、章节名称）使用难度颜色
                cell.fill = difficulty_fill
    
    # 设置列宽
    single_col_widths = [6, 10, 25, 8, 40, 20, 20, 20, 20, 6, 40, 8]
    for col_num, width in enumerate(single_col_widths, 1):
        ws_single.column_dimensions[get_column_letter(col_num)].width = width
    
    # ========== 多选题工作表 ==========
    ws_multiple = wb.create_sheet(title="多选题")
    
    # 设置表头
    multiple_headers = ['题号', '章节', '章节名称', '页码', '题目内容', '选项A', '选项B', '选项C', '选项D', '答案', '解析', '难度']
    ws_multiple.append(multiple_headers)
    
    # 设置表头样式
    for col_num, header in enumerate(multiple_headers, 1):
        cell = ws_multiple.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 填充数据
    for idx, question in enumerate(data['multiple_choice'], 1):
        options = question.get('options', [])
        answer = question.get('answer', [])
        if isinstance(answer, list):
            answer = ','.join(answer)
        
        row_data = [
            idx,
            question.get('chapter', ''),
            question.get('chapter_title', ''),
            question.get('page', ''),
            question.get('question', ''),
            options[0].replace('A. ', '') if len(options) > 0 else '',
            options[1].replace('B. ', '') if len(options) > 1 else '',
            options[2].replace('C. ', '') if len(options) > 2 else '',
            options[3].replace('D. ', '') if len(options) > 3 else '',
            answer,
            question.get('analysis', ''),
            question.get('difficulty', '中等')
        ]
        
        ws_multiple.append(row_data)
        
        # 设置内容样式
        row_num = idx + 1
        difficulty = question.get('difficulty', '中等')
        difficulty_fill = difficulty_colors.get(difficulty, difficulty_colors['中等'])
        
        for col_num in range(1, len(row_data) + 1):
            cell = ws_multiple.cell(row=row_num, column=col_num)
            cell.font = content_font
            cell.alignment = content_alignment
            cell.border = thin_border
            
            if col_num <= 3:
                cell.fill = difficulty_fill
    
    # 设置列宽
    for col_num, width in enumerate(single_col_widths, 1):
        ws_multiple.column_dimensions[get_column_letter(col_num)].width = width
    
    # ========== 简答题工作表 ==========
    ws_short = wb.create_sheet(title="简答题")
    
    # 设置表头
    short_headers = ['题号', '章节', '章节名称', '页码', '题目内容', '答案', '解析', '难度']
    ws_short.append(short_headers)
    
    # 设置表头样式
    for col_num, header in enumerate(short_headers, 1):
        cell = ws_short.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 填充数据
    for idx, question in enumerate(data['short_answer'], 1):
        row_data = [
            idx,
            question.get('chapter', ''),
            question.get('chapter_title', ''),
            question.get('page', ''),
            question.get('question', ''),
            question.get('answer', ''),
            question.get('analysis', ''),
            question.get('difficulty', '中等')
        ]
        
        ws_short.append(row_data)
        
        # 设置内容样式
        row_num = idx + 1
        difficulty = question.get('difficulty', '中等')
        difficulty_fill = difficulty_colors.get(difficulty, difficulty_colors['中等'])
        
        for col_num in range(1, len(row_data) + 1):
            cell = ws_short.cell(row=row_num, column=col_num)
            cell.font = content_font
            cell.alignment = content_alignment
            cell.border = thin_border
            
            if col_num <= 3:
                cell.fill = difficulty_fill
    
    # 设置列宽
    short_col_widths = [6, 10, 25, 8, 50, 50, 50, 8]
    for col_num, width in enumerate(short_col_widths, 1):
        ws_short.column_dimensions[get_column_letter(col_num)].width = width
    
    # ========== 统计信息工作表 ==========
    ws_stats = wb.create_sheet(title="统计信息")
    
    # 添加统计信息
    stats_data = [
        ['试题统计信息'],
        [''],
        ['题目类型', '数量'],
        ['单选题', len(data['single_choice'])],
        ['多选题', len(data['multiple_choice'])],
        ['简答题', len(data['short_answer'])],
        ['总计', sum(len(v) for v in data.values())],
        [''],
        ['难度分布'],
        ['难度', '单选题', '多选题', '简答题', '合计'],
        ['基础',
         sum(1 for q in data['single_choice'] if q.get('difficulty') == '基础'),
         sum(1 for q in data['multiple_choice'] if q.get('difficulty') == '基础'),
         sum(1 for q in data['short_answer'] if q.get('difficulty') == '基础'),
         sum(1 for q in data['single_choice'] if q.get('difficulty') == '基础') +
         sum(1 for q in data['multiple_choice'] if q.get('difficulty') == '基础') +
         sum(1 for q in data['short_answer'] if q.get('difficulty') == '基础')],
        ['中等',
         sum(1 for q in data['single_choice'] if q.get('difficulty') == '中等'),
         sum(1 for q in data['multiple_choice'] if q.get('difficulty') == '中等'),
         sum(1 for q in data['short_answer'] if q.get('difficulty') == '中等'),
         sum(1 for q in data['single_choice'] if q.get('difficulty') == '中等') +
         sum(1 for q in data['multiple_choice'] if q.get('difficulty') == '中等') +
         sum(1 for q in data['short_answer'] if q.get('difficulty') == '中等')],
        ['困难',
         sum(1 for q in data['single_choice'] if q.get('difficulty') == '困难'),
         sum(1 for q in data['multiple_choice'] if q.get('difficulty') == '困难'),
         sum(1 for q in data['short_answer'] if q.get('difficulty') == '困难'),
         sum(1 for q in data['single_choice'] if q.get('difficulty') == '困难') +
         sum(1 for q in data['multiple_choice'] if q.get('difficulty') == '困难') +
         sum(1 for q in data['short_answer'] if q.get('difficulty') == '困难')],
        [''],
        ['教材信息'],
        ['书名', '《火电厂热工自动控制技术及应用》'],
        ['作者', '刘禾、白焰、李新利'],
        ['出版社', '中国电力出版社'],
        [''],
        ['颜色说明'],
        ['绿色', '基础难度'],
        ['橙色', '中等难度'],
        ['红色', '困难难度'],
    ]
    
    for row_data in stats_data:
        ws_stats.append(row_data)
    
    # 设置统计表样式
    ws_stats.column_dimensions['A'].width = 15
    ws_stats.column_dimensions['B'].width = 15
    ws_stats.column_dimensions['C'].width = 15
    ws_stats.column_dimensions['D'].width = 15
    ws_stats.column_dimensions['E'].width = 15
    
    # 为标题行设置样式
    ws_stats['A1'].font = Font(name='微软雅黑', size=14, bold=True)
    ws_stats['A3'].font = Font(name='微软雅黑', size=12, bold=True)
    ws_stats['B3'].font = Font(name='微软雅黑', size=12, bold=True)
    ws_stats['A9'].font = Font(name='微软雅黑', size=12, bold=True)
    ws_stats['A10'].font = Font(name='微软雅黑', size=11, bold=True)
    ws_stats['B10'].font = Font(name='微软雅黑', size=11, bold=True)
    ws_stats['C10'].font = Font(name='微软雅黑', size=11, bold=True)
    ws_stats['D10'].font = Font(name='微软雅黑', size=11, bold=True)
    ws_stats['E10'].font = Font(name='微软雅黑', size=11, bold=True)
    
    # 为难度行设置颜色
    ws_stats['A11'].fill = difficulty_colors['基础']
    ws_stats['A12'].fill = difficulty_colors['中等']
    ws_stats['A13'].fill = difficulty_colors['困难']
    
    # 保存文件
    wb.save(output_file)
    print(f"Excel文件已创建: {output_file}")
    
    return output_file


if __name__ == "__main__":
    questions_file = "/tmp/generated_questions.json"
    output_file = "/tmp/热工自动控制试题集.xlsx"
    
    # 检查openpyxl是否安装
    try:
        from openpyxl import Workbook
    except ImportError:
        print("正在安装openpyxl...")
        os.system("pip install openpyxl -q")
    
    create_questions_excel(questions_file, output_file)
    
    # 显示文件信息
    file_size = os.path.getsize(output_file)
    print(f"文件大小: {file_size / 1024:.2f} KB")
